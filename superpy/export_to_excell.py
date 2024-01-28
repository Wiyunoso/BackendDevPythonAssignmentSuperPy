import pandas
import openpyxl
from openpyxl.utils import get_column_letter
from rich.console import Console

console = Console()


def export_to_excell(file):
    if file == "bought":
        read_file = pandas.read_csv(r"./csv_files/bought.csv")
        read_file.to_excel(r"./excell_files/bought.xlsx", index=None, header=True)
    elif file == "inventory":
        read_file = pandas.read_csv(r"./csv_files/inventory.csv")
        read_file.to_excel(r"./excell_files/inventory.xlsx", index=None, header=True)
    elif file == "sold":
        read_file = pandas.read_csv(r"./csv_files/sold.csv")
        read_file.to_excel(r"./excell_files/sold.xlsx", index=None, header=True)
    elif file == "expired":
        read_file = pandas.read_csv(r"./csv_files/expired.csv")
        read_file.to_excel(r"./excell_files/expired.xlsx", index=None, header=True)
    adjust_excell_columns_width("./excell_files/" + file + ".xlsx")
    console.print(
        "The "
        + file
        + ".csv-file has been exported as Excell-file in the folder 'excell_files'.",
        style="green",
    )


def adjust_excell_columns_width(file):
    wb = openpyxl.load_workbook(filename=file)
    worksheet = wb.active
    for idx, col in enumerate(worksheet.columns, 1):
        worksheet.column_dimensions[get_column_letter(idx)].auto_size = True
    wb.save(file)
